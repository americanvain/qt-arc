
from PySide2.QtWidgets import QApplication, QMessageBox,QLineEdit
from PySide2.QtUiTools import QUiLoader
from PySide2 import QtCore
from openpyxl import Workbook,load_workbook
import PySide2
# import json



class excel_data:
    name=''
    id_name=''
    pos_excel=''
    pos_qt=''
    value_true=None
    default_value=''
    cal_fun=''

    def __init__(self,name_,pos_qt_,pos_excel_,default_value_):
        self.name=name_.text()
        self.pos_excel=pos_excel_
        self.pos_qt=pos_qt_
        self.default_value=default_value_
        if (type(self.pos_qt)==PySide2.QtWidgets.QLineEdit):
            pos_qt_.setPlaceholderText(self.default_value)
            pos_qt_.setText(self.default_value)
            self.value_true=float(self.pos_qt.text())
        elif (type(self.pos_qt)==PySide2.QtWidgets.QComboBox):
            pos_qt_.setCurrentIndex(self.default_value)
            self.value_true=self.pos_qt.currentText()
            

class Stats:

    name_list=[]
    value_list=[]
    d={}
    excel_pos=[]
    
    #混凝土强度等级
    def __init__(self):
        # 从文件中加载UI定义

        # 从 UI 定义中动态 创建一个相应的窗口对象
        # 注意：里面的控件对象也成为窗口对象的属性了
        # 比如 self.ui.button , self.ui.textEdit
        self.ui = QUiLoader().load('arcuifile.ui')
        self.ui.pushButton.clicked.connect(self.handleCalc)
        self.ui.pushButton_2.clicked.connect(self.handleExcelXieru)
        self.ui.pushButton_3.clicked.connect(self.handleExcelyulan)
        # self.ui_init()
        # self.data_init()

    def ui_init(self):
        self.ui.comboBox.setCurrentIndex(3)
        self.ui.lineEdit.setText("1")
        self.ui.lineEdit_2.setText('360')
        self.ui.lineEdit_3.setText('360')
        self.ui.lineEdit_4.setText('200000')
        self.ui.lineEdit_5.setText('300')
        self.ui.lineEdit_6.setText('1800')
        self.ui.lineEdit_7.setText('1320')
        self.ui.lineEdit_8.setText('390')
        self.ui.lineEdit_9.setText('195000')


    def data_init(self):
        #混凝土强度等级
        混凝土强度等级=excel_data(self.ui.label,self.ui.comboBox,"F3",3)
        self.name_list.append(self.ui.label.text())
        self.value_list.append(self.ui.comboBox.currentText())
        self.excel_pos.append("F3")
        #张拉控制强度系数
        张拉控制强度系数=excel_data(self.ui.label_2,self.ui.lineEdit,"F9",1)
        self.name_list.append(self.ui.label_2.text())
        self.value_list.append(int(self.ui.lineEdit.text()))
        self.excel_pos.append("F9")
        #HRB400(φ)钢筋抗拉强度设计值fy=
        张拉控制强度系数=excel_data(self.ui.label_2,self.ui.lineEdit,"F9",1)
        self.name_list.append(self.ui.label_3.text())
        self.value_list.append(int(self.ui.lineEdit_2.text()))
        self.excel_pos.append("F13")
        #HRB400(φ)钢筋抗压强度设计值f′y=
        self.name_list.append(self.ui.label_4.text())
        self.value_list.append(int(self.ui.lineEdit_3.text()))
        self.excel_pos.append("F14")
        #HRB400(φ)钢筋、HRB335(φ)钢筋弹模Es=
        self.name_list.append(self.ui.label_5.text())
        self.value_list.append(int(self.ui.lineEdit_4.text()))
        self.excel_pos.append("F15")
        #HRB335φ钢筋抗拉强度设计值fpy=
        self.name_list.append(self.ui.label_6.text())
        self.value_list.append(int(self.ui.lineEdit_5.text()))
        self.excel_pos.append("F16")
        #预应力钢绞线强度标准值fptk=
        self.name_list.append(self.ui.label_7.text())
        self.value_list.append(int(self.ui.lineEdit_6.text()))
        self.excel_pos.append("F17")
        #抗拉强度设计值fpy=
        self.name_list.append(self.ui.label_8.text())
        self.value_list.append(int(self.ui.lineEdit_7.text()))
        self.excel_pos.append("F18")
        #抗压强度设计值f′py=
        self.name_list.append(self.ui.label_9.text())
        self.value_list.append(int(self.ui.lineEdit_8.text()))
        self.excel_pos.append("F19")
        #钢绞线弹模E′p=
        self.name_list.append(self.ui.label_10.text())
        self.value_list.append(int(self.ui.lineEdit_9.text()))
        self.excel_pos.append("F20")

        for i in range(len(self.name_list)):
            self.d[self.name_list[i]]=self.value_list[i]


    def handleCalc(self):
        info = self.ui.comboBox.currentText()
        print(type(info))

    def handleExcelXieru(self):
        hunningtuqiangdu = self.ui.comboBox.currentText()
        zhanglakongzhixishu=float(self.ui.lineEdit.text())
        fy=float(self.ui.lineEdit_2.text())
        f_y=float(self.ui.lineEdit_3.text())
        Es=float(self.ui.lineEdit_4.text())
        fpy=float(self.ui.lineEdit_5.text())
        fptk=float(self.ui.lineEdit_6.text())
        fpy2=float(self.ui.lineEdit_7.text())
        f_py=float(self.ui.lineEdit_8.text())
        E_p=float(self.ui.lineEdit_9.text())

        file = 'newmyexcel.xlsx'
        wb = load_workbook(file)
        # print(wb.sheetnames)
        sheet = wb["part1-基本参数"]
        sheet["F3"].value=self.ui.comboBox.currentText()
        sheet["F9"].value=float(self.ui.lineEdit.text())
        sheet["F13"].value=float(self.ui.lineEdit_2.text())
        sheet["F14"].value=float(self.ui.lineEdit_3.text())
        sheet["F15"].value=float(self.ui.lineEdit_4.text())
        sheet["F16"].value=float(self.ui.lineEdit_5.text())
        sheet["F17"].value=float(self.ui.lineEdit_6.text())
        sheet["F18"].value=float(self.ui.lineEdit_7.text())
        sheet["F19"].value=float(self.ui.lineEdit_8.text())
        sheet["F20"].value=float(self.ui.lineEdit_9.text())

        sheet["E25"].value=float(self.ui.lineEdit_10.text())
        sheet["E26"].value=float(self.ui.lineEdit_11.text())
        sheet["E27"].value=float(self.ui.lineEdit_12.text())
        sheet["E30"].value=float(self.ui.lineEdit_13.text())
        sheet["E32"].value=float(self.ui.lineEdit_14.text())
        sheet["E33"].value=float(self.ui.lineEdit_15.text())
        sheet["E35"].value=float(self.ui.lineEdit_16.text())
        sheet["E36"].value=float(self.ui.lineEdit_17.text())
        sheet["E37"].value=float(self.ui.lineEdit_18.text())
        sheet["E38"].value=float(self.ui.lineEdit_19.text())
        sheet["E41"].value=float(self.ui.lineEdit_20.text())
        sheet["E43"].value=float(self.ui.lineEdit_21.text())
        # sheet[""].value=float(self.ui.lineEdit_22.text())
        # sheet[""].value=float(self.ui.lineEdit_23.text())
        sheet["E45"].value=float(self.ui.lineEdit_24.text())
        sheet["E47"].value=float(self.ui.lineEdit_25.text())
        sheet["E49"].value=float(self.ui.lineEdit_26.text())
        sheet["E51"].value=float(self.ui.lineEdit_27.text())
        sheet["E53"].value=float(self.ui.lineEdit_28.text())
        sheet["E55"].value=float(self.ui.lineEdit_29.text())
        sheet["E56"].value=float(self.ui.lineEdit_30.text())
        sheet["E59"].value=float(self.ui.lineEdit_31.text())
        sheet["D66"].value=float(self.ui.lineEdit_32.text())
        sheet["F66"].value=float(self.ui.lineEdit_33.text())
        sheet["D67"].value=float(self.ui.lineEdit_34.text())
        sheet["F67"].value=float(self.ui.lineEdit_35.text())
        sheet["E72"].value=float(self.ui.lineEdit_36.text())
        sheet["E73"].value=self.ui.comboBox_2.currentText()


        wb.save('newtest2.xlsx')
        print("finished")

    def handleExcelyulan(self):
        hunningtuqiangdu = self.ui.comboBox.currentText()
        zhanglakongzhixishu=self.ui.lineEdit.text()
        fy=self.ui.lineEdit_2.text()
        f_y=self.ui.lineEdit_3.text()
        Es=self.ui.lineEdit_4.text()
        fpy=self.ui.lineEdit_5.text()
        fptk=self.ui.lineEdit_6.text()
        fpy2=self.ui.lineEdit_7.text()
        E_p=self.ui.lineEdit_8.text()
        temp =int(zhanglakongzhixishu)
        print(self.ui.label.text(),hunningtuqiangdu)
        print(self.ui.label_2.text(),type(temp))


        
    # def excel_data_init(self):
    #     self.ui.lineEdit.text()


if __name__ == "__main__":
    QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_ShareOpenGLContexts)
    app = QApplication([])
    stats = Stats()
    stats.ui.show()
    app.exec_()